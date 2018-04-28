import Myoption
import importlib
import numpy as np
from tensorflow.python.platform import gfile
from data import *
from evaluate import evaluate
from tqdm import tqdm
import customfunc
import nnUtils
FLAGS = tf.app.flags.FLAGS
# tf.logging.set_verbosity(FLAGS.log)

##Writer에 작성 할 내용을 정의하는 함수
def add_summaries(scalar_list=[], activation_list=[], grad_list=[], var_list=[],Wbin_list=[],Wfluc_list=[],Drift_step=[],Drift_value=[]):
    #Name:
    #Quantized weight=Real_Binarized_weight
    #Target_resistance=Binarized_weight
    #Written_resistance=Fluctuated_weight
    #Read_weight=Real_Fluctuated_weight
    # 정확도
    for var in scalar_list:
        if var != None:
            tf.summary.scalar((var.op.name).split('/')[-1]+'/training_real', var)
    if FLAGS.summary2==True:
        # activation histogram
        for activation in activation_list:
            if activation != None:
                ful=activation.op.name
                tf.summary.histogram(ful.split('/')[0]+'/2/'+'activations', activation)
        #Gradients histogram
        for grad, var in grad_list:
            if grad != None:
                tf.summary.histogram(var.op.name + '/gradients', grad)

        zip_list=[]
        Wquan_list = tf.get_collection(key='Quantized_Weight')
        R_Wfluc_list = tf.get_collection(key='Read_Weight')
        # Drift_step,Drift_value는 여기서 처리하기에는 예외가 많아서 nnUtils에서 drift꺼지더라도 직접 []을 추가해주기로 함
        # 예를 들어서 Drift1=False, Drift2=True면 Drift_step이 Wbin_list보다 짧게 나옴, 아래의 코드로 처리불가
        length=len(var_list)
        zip_list.append([[]] * length) if var_list==[] else zip_list.append(var_list) #이 코드는 그냥 아예 리스트가 안들어왔을 때를 위해
        zip_list.append([[]] * length) if Wbin_list==[] else zip_list.append(Wbin_list)
        zip_list.append([[]] * length) if Wfluc_list==[] else zip_list.append(Wfluc_list)
        zip_list.append([[]] * length) if R_Wfluc_list == [] else zip_list.append(R_Wfluc_list)
        zip_list.append([[]] * length) if Wquan_list==[] else zip_list.append(Wquan_list)
        zip_list.append(Drift_step)
        zip_list.append(Drift_value)
        for W,Wbin,Wfluc,Real_Wfluc,Real_Wbin,Step,Value in zip(*zip_list):
            # 1:이름 뽑아내기, 물론 여기서 새로 Fluctuated, Binarized 식으로 쓰는게 더 직관적이지만 기존의 텐서에서 이름 뽑아내는 명령어와
            # 2:split 함수 써보기 위해서 그냥 이렇게 한다.
            # 3:이 부분을 위에서 gradients 처리한 것처럼, W,Wbin,Wfluc 따로따로해줘도 된다, 어차피  똑같은 방식들이라, 길이 줄이려고 이렇게 한건데
            # if문이 많이 들어가서 피차일반의 결과인 듯.
            # Histogram and name
            if W != []:
                name_layer = (W.op.name).split('/')[0]  # W가 Wbin,Wfluc으로 바뀌어도 어차피 / 기준 첫번째는 레이어 이름이라 결과는 똑같다.
                tf.summary.histogram(name_layer + '/0/' + "Original_weight", W)
            else:
                print("No W_list")
            if Wbin != []:
                tf.summary.histogram(name_layer + '/1/' + 'Target_resistance', Wbin)
            else:
                print("No Wbin_list")
            if Wfluc!=[]:
                tf.summary.histogram(name_layer + '/2/' + 'Written_resistance', Wfluc)
            else:
                print("No Wfluc_list")
            if Real_Wfluc != []:
                tf.summary.histogram(name_layer + '/3/' + 'Read_weight', Real_Wfluc)
            else:
                print("No Wbin_list")
            if Real_Wbin != []:
                tf.summary.histogram(name_layer + '/4/' + 'Quantized_weight', Real_Wbin)
            else:
                print("No Wbin_list")
            # Kernel - mainly for CNN
            sz = W.get_shape().as_list()
            if len(sz) == 4 and sz[2] == 3:
                kernels = tf.transpose(W, [3, 0, 1, 2])
                tf.summary.image(W.op.name + '/kernels', group_batch_images(kernels), max_outputs=1)

            # Drift ratio part: all elements in the 'Value' should bigger than 0
            if Value!=[] and Value.get_shape().as_list()[0] != 0:
                assert_op = tf.Assert(tf.reduce_min(Value) >= 0, [Value])
                with tf.control_dependencies([assert_op]):
                    Value=tf.identity(Value)

            # Some weights in the layer
            num_weights = 5
            index_history=[]
            for i in range(num_weights):
                index = [np.random.randint(j) for j in W.get_shape().as_list()]
                index_str = '/' + str(index).replace(', ','_')[1:-1]
                index_history.append(index)
                if W != []:
                    tf.summary.scalar(name_layer+ index_str + '/0/' + 'Original_weight', W[index])
                if Wbin != []:
                    tf.summary.scalar(name_layer+ index_str + '/1/' + 'Target_resistance', Wbin[index])
                if Wfluc != []:
                    tf.summary.scalar(name_layer+ index_str + '/2/' + 'Written_resistance', Wfluc[index])
                if Real_Wfluc != []:
                    tf.summary.scalar(name_layer + index_str + '/3/' + 'Read_weight', Real_Wfluc[index])
                if Real_Wbin != []:
                    tf.summary.scalar(name_layer + index_str + '/4/' + 'Quantized_weight', Real_Wbin[index])
                if Value!=[] and Value.get_shape().as_list()[0] != 0:
                    tf.summary.scalar(name_layer+ index_str + '/5/' + 'dvalue', Value[index])

            file = open(FLAGS.checkpoint_dir + "/save_model.py", "a")
            print("'''",name_layer,' : ',index_history,"'''",file=file)
            file.close()

            # Calculate the ratio (Drifted weights num)/(all weights num)
            if Step!=[] and Step.get_shape().as_list()[0] != 0:
                num=Step.get_shape().num_elements()
                weights_keeping1  = tf.cast(Step>20,dtype=tf.float32)
                weights_keeping2 = tf.cast(Step > 120, dtype=tf.float32)
                weights_keeping3 = tf.cast(Step > 840, dtype=tf.float32)

                weights_keeping_num1 =tf.reduce_sum(weights_keeping1)
                weights_keeping_num2 = tf.reduce_sum(weights_keeping2)
                weights_keeping_num3 = tf.reduce_sum(weights_keeping3)

                weights_keeping_num_reset1 = tf.reduce_sum(weights_keeping1 * tf.cast(Wbin >= 0, dtype=tf.float32))
                weights_keeping_num_reset2 = tf.reduce_sum(weights_keeping2 * tf.cast(Wbin >= 0, dtype=tf.float32))
                weights_keeping_num_reset3 = tf.reduce_sum(weights_keeping3 * tf.cast(Wbin >= 0, dtype=tf.float32))

                ratio11 = weights_keeping_num1 / num
                ratio12 = weights_keeping_num2 / num
                ratio13 = weights_keeping_num3 / num
                ratio21 = weights_keeping_num_reset1 / tf.reduce_sum(tf.cast(Wbin>=0,dtype=tf.float32))
                ratio22 = weights_keeping_num_reset2 / tf.reduce_sum(tf.cast(Wbin>=0,dtype=tf.float32))
                ratio23 = weights_keeping_num_reset3 / tf.reduce_sum(tf.cast(Wbin>=0,dtype=tf.float32))

                tf.summary.scalar(name_layer + "/Ratio_Keeping/1_125", ratio11)
                tf.summary.scalar(name_layer + "/Ratio_Keeping/1_205", ratio12)
                tf.summary.scalar(name_layer + "/Ratio_Keeping/1_300", ratio13)
                tf.summary.scalar(name_layer + "/Ratio_Drifted/1_125", ratio21)
                tf.summary.scalar(name_layer + "/Ratio_Drifted/1_205", ratio22)
                tf.summary.scalar(name_layer + "/Ratio_Drifted/1_300", ratio23)
    #tf.summary.scalar(activation.op.name + '/sparsity', tf.nn.zero_fraction(activation))

if FLAGS.dataset=='cifar10':
    steps=round(50000/FLAGS.batch_size)*20
else:
    steps=round(55000/FLAGS.batch_size)*20
##LR을 decay시켜주는 함수
def learning_rate_decay_fn(learning_rate, global_step,decay_steps=steps):
    print("learning_rate_decay_fn is executed!")
    return tf.train.exponential_decay(
      learning_rate,
      global_step,
      decay_steps=decay_steps,
      decay_rate=0.8,
      staircase=True)
def quantizeGrads(Grad_and_vars,target_level=FLAGS.W_target_level):
    if target_level <= 256*256:
        grads = []
        for grad_and_vars in Grad_and_vars:
            grads.append([nnUtils.quantize_G(grad_and_vars[0],target_level), grad_and_vars[1]])
        return grads
    return Grad_and_vars
## model을 data로 training 시켜주는 함수
def train(model, data,
          batch_size=128,
          learning_rate=FLAGS.learning_rate,
          Weight_decay=FLAGS.Weight_decay,
          log_dir='./log',
          checkpoint_dir='./checkpoint',
          num_epochs=-1):

    with tf.name_scope('data'):
        x, yt = data.next_batch(batch_size)
    global_step =  tf.get_variable('global_step', shape=[],dtype=tf.int64,
                         initializer=tf.constant_initializer(0),
                         trainable=False)
    tf.add_to_collection("Step",global_step)  #Evaluate에서 Drift효과 끄기 위해 구분점역할을 한다.
    y = model(x, is_training=True)

    # Define loss and optimizer
    with tf.name_scope('objective'):
        yt_one=tf.one_hot(yt,10)
        loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(labels=yt_one, logits=y), name="loss")
        if Weight_decay:
            loss=loss+tf.reduce_sum(Myoption.WEIGHT_DECAY_FACTOR*tf.stack([tf.nn.l2_loss(i) for i in tf.get_collection('Original_Weight', scope='L')]))
        accuracy=tf.reduce_mean(tf.cast(tf.equal(tf.arg_max(yt_one,1), tf.argmax(y, axis=1)),dtype=tf.float32),name="accuracy")
        # *이런 함수도 있어요.
        #accuracy = tf.reduce_mean(tf.cast(tf.nn.in_top_k(y, yt, 1), tf.float32))

    vars_train=tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L22')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L23')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L24')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L25') if FLAGS.Fine_tuning else None
    # * 원래는 이런 optimizer 썼었다,  근데 그 아래가 기능이 훨씬 많아서 갈아탔다.
    # lambda learning_rate: tf.train.MomentumOptimizer(learning_rate, momentum=0.9)
    # opt = tf.contrib.layers.optimize_loss(loss, global_step, learning_rate, optimizer='SGD',
    #                                       gradient_noise_scale=None, gradient_multipliers=None,
    #                                       clip_gradients=None, # moving_average_decay=0.9,
    #                                        update_ops=None, variables=vars_train, name=None,
    #                                       learning_rate_decay_fn=learning_rate_decay_fn)
    # # * 이런 식으로 gradient  뽑아서  수정가능
    optimizer=tf.train.GradientDescentOptimizer(1)
    grads = optimizer.compute_gradients(loss)
    gradTrainBatch_quantize = quantizeGrads(grads,FLAGS.W_target_level)
    opt = optimizer.apply_gradients(gradTrainBatch_quantize, global_step=global_step)

    print("Definite Moving Average...")
    ema = tf.train.ExponentialMovingAverage(Myoption.MOVING_AVERAGE_DECAY, global_step, name='average')
    ema_op = ema.apply([loss, accuracy]+tf.trainable_variables())
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, ema_op)

    loss_avg = ema.average(loss)
    tf.summary.scalar('loss/training', loss_avg)
    accuracy_avg = ema.average(accuracy)
    tf.summary.scalar('accuracy/training', accuracy_avg)
    check_loss = tf.check_numerics(loss, 'model diverged: loss->nan')
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, check_loss)
    list_W = tf.get_collection('Original_Weight', scope='L')
    list_Wbin = tf.get_collection('Binarized_Weight', scope='L')
    list_Wfluc = tf.get_collection('Fluctuated_Weight', scope='L')
    list_Drift_step = tf.get_collection('Drift_step',scope='L')
    list_Drift_value= tf.get_collection('Drift_value',scope='L')
    list_pre_Wbin = tf.get_collection('pre_Wbin', scope='L')
    list_pre_Wfluc = tf.get_collection('pre_Wfluc', scope='L')
    list_pre_Wbin_op = tf.get_collection('pre_Wbin_update_op', scope='L')
    list_pre_Wfluc_op = tf.get_collection('pre_Wfluc_update_op', scope='L')
    clip_op_list=[]
    for ww in list_W:
        clip_op=tf.assign(ww,nnUtils.clip(ww,FLAGS.W_target_level))
        clip_op_list+=[clip_op]
    updates_collection = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
    # updates_collection은 여러 operation으로 이루어진 리스트고, train_op는 그 operation들을 실행하는 operation이다.
    with tf.control_dependencies([opt]):
        train_op = tf.group(*updates_collection)
    print("Make summary for writer...")

    if FLAGS.summary:
        add_summaries(scalar_list=[accuracy, loss],activation_list=tf.get_collection(tf.GraphKeys.ACTIVATIONS),
            var_list=list_W,Wbin_list=list_Wbin,Wfluc_list=list_Wfluc,Drift_step=list_Drift_step,Drift_value=list_Drift_value)
            #grad_list=grads)
    summary_op = tf.summary.merge_all()

    print("Open Session...")
    # Configure options for session
    gpu_options = tf.GPUOptions(allow_growth=True,per_process_gpu_memory_fraction=0.8)
    sess = tf.InteractiveSession(
        config=tf.ConfigProto(
            log_device_placement=False,
            allow_soft_placement=True, gpu_options=gpu_options))
    sess.run(tf.global_variables_initializer())
    for ww in list_W:
         bbb=sess.run(tf.assign(ww,nnUtils.quantize_W(ww,FLAGS.W_target_level)))
    if FLAGS.Load_checkpoint:
        call_list=tf.get_collection(tf.GraphKeys.VARIABLES)
        print("******We will restore:",call_list)
        saver = tf.train.Saver(max_to_keep=1,var_list=call_list)
        ckpt = tf.train.get_checkpoint_state(Myoption.load_checkpoint_path)
        if ckpt and ckpt.model_checkpoint_path:
            # Restores from checkpoint
            saver.restore(sess, ckpt.model_checkpoint_path)
        else:
            print('No checkpoint file found')
    saver = tf.train.Saver(max_to_keep=1)
    saver_best= tf.train.Saver(max_to_keep=1)

    summary_writer = tf.summary.FileWriter(log_dir, graph=sess.graph)
    coord = tf.train.Coordinator()
    threads = tf.train.start_queue_runners(sess=sess, coord=coord)
    best_acc = 0
    patience = 0
    num_batches = int(data.size[0] / batch_size)

    print("Check the collections...")
    print("list_W:\n",list_W,'\nNum:',len(list_W))
    print("Activations:\n", tf.get_collection(tf.GraphKeys.ACTIVATIONS), '\nNum:', len(tf.get_collection(tf.GraphKeys.ACTIVATIONS)))

    file = open(FLAGS.checkpoint_dir + "/save_model.py", "a")
    customfunc.magic_print(
        'We start training..num of trainable param: %d' % customfunc.count_params(tf.trainable_variables()),file=file)
    customfunc.magic_print(
        "Drift1 : ",FLAGS.Drift1,"\nDrift2 : ",FLAGS.Drift2,"\nVariation : ",FLAGS.Variation,file=file)
    customfunc.magic_print(
        "LR : ", FLAGS.learning_rate, "\nbatch_size : ", FLAGS.batch_size, "\nDataset : ", FLAGS.dataset, file=file)
    for i in range(num_epochs):
        if len(Myoption.LR_schedule)/2:
            if i == Myoption.LR_schedule[0]:
                Myoption.LR_schedule.pop(0)
                LR_new = Myoption.LR_schedule.pop(0)
                if LR_new == 0:
                    print('Optimization Ended!')
                    exit(0)
                LR_old = sess.run(Myoption.LR)
                sess.run(Myoption.LR.assign(LR_new))
                print('lr: %f -> %f' % (LR_old, LR_new))
        customfunc.magic_print('Started epoch %d' % (i + 1),file=file)
        count_num=np.array([0,0,0,0,0,0,0,0,0,0])
        for j in tqdm(range(num_batches)):
            list_run = sess.run(list_Wbin+list_Wfluc+[train_op, loss]+[y,yt])
            sess.run(clip_op_list)
            # 각 iteration에서 train_op를 통해 업데이트를 하기 전에 list_Wbin,Wfluc에 있는 var들의 값을 save for next batch
            unique_elements,elements_counts=np.unique(list_run[-1],return_counts=True)
            num_set=dict(zip(unique_elements,elements_counts))
            #ii라는 숫자가 dictionary에 들어있다면 카운트에 더해준다.
            for ii in range(10):
                if num_set.__contains__(ii):
                    count_num[ii]=count_num[ii]+num_set[ii]
            if FLAGS.Variation:
                for index, value in enumerate(list_run[0:len(list_Wbin)]):
                    sess.run(list_pre_Wbin_op[index],{list_pre_Wbin[index]:value})
                for index, value in enumerate(list_run[len(list_Wbin):len(list_Wbin + list_Wfluc)]):
                    sess.run(list_pre_Wfluc_op[index],{list_pre_Wfluc[index]:value})
            if j%100==0:
                summary_writer.add_summary(sess.run(summary_op), global_step=sess.run(global_step))
        # print("Output:", list_run[-2][:10])
        # print("Softmax:", sess.run(tf.nn.softmax(list_run[-2][:10])))
        # print("loss:", sess.run(tf.nn.softmax_cross_entropy_with_logits(labels=tf.one_hot(list_run[-1],10),logits=list_run[-2]))[:10])
        step, acc_value, loss_value, summary = sess.run([global_step, accuracy_avg, loss_avg, summary_op])
        """
        20171204:avg기능을 빼고 코드를 돌려보니까 training set에서의 정확도가 98퍼가 되었다가 96퍼가 되었다가 왔다갔다한다.
        이유: 위에서 accuracy 를 돌릴 때 데이터가 64개밖에 안쓰인다, 그래서 정확도의 격차가 좀 있었던 것
        """
        customfunc.magic_print(
            ["%d : " % i + str(count_num[i]) for i in range(10)], " Totral num: ", count_num.sum(),file=file)
        customfunc.magic_print('Training - Accuracy: %.3f' % acc_value, '  Loss:%.3f' % loss_value,file=file)
        saver.save(sess, save_path=checkpoint_dir + '/model.ckpt', global_step=global_step)

        test_acc, test_loss = evaluate(model, FLAGS.dataset,checkpoint_dir=checkpoint_dir)# log_dir=log_dir)
        customfunc.magic_print('Test     - Accuracy: %.3f' % test_acc, '  Loss:%.3f' % test_loss,file=file)

        if best_acc<test_acc:
            best_acc=test_acc
            patience=0
            saver_best.save(sess, save_path=checkpoint_dir + '/best_model.ckpt', global_step=global_step,latest_filename="best_checkpoint")
        elif best_acc >= test_acc:
            patience += 1
            if patience > 20:
                customfunc.magic_print("Stop this training at epoch"+str(i+1)+", because accuracy may be saturated",file=file)
                from openpyxl import Workbook
                from openpyxl import load_workbook
                file_name = 'Accuracy_log'
                if gfile.Exists(file_name+'.xlsx'):
                    wb=load_workbook(filename=file_name+'.xlsx')
                    ws=wb[file_name]
                else:
                    wb=Workbook()
                    ws=wb.active
                    ws.title=file_name
                row=10 if FLAGS.dataset=='cifar10' else 2
                ws.cell(column=FLAGS.target_level+1, row=row, value=best_acc)
                ws.cell(column=FLAGS.target_level+1, row=row + 1, value=test_acc)
                ws.cell(column=FLAGS.target_level+1, row=row + 2, value=acc_value)
                ws.cell(column=FLAGS.target_level+1, row=row + 3, value=i)
                wb.save(filename=file_name+'.xlsx')
                break
        customfunc.magic_print('Best     - Accuracy: %.3f(patience=%d)' % (best_acc,patience),file=file)
        summary_out = tf.Summary()
        summary_out.ParseFromString(summary)
        summary_out.value.add(tag='accuracy/test', simple_value=test_acc)
        summary_out.value.add(tag='loss/test', simple_value=test_loss)
        summary_writer.add_summary(summary_out, step)
        summary_writer.flush()


    # When done, ask the threads to stop.
    file.close()
    coord.request_stop()
    coord.join(threads)
    coord.clear_stop()
    summary_writer.close()
"""
설명2:
1)what is the argv
Argv in Python
The list of command line arguments passed to a Python script. argv[0] is the script name (it is operating system 
dependent whether this is a full pathname or not). If the command was executed using the -c command line option 
to the interpreter, argv[0] is set to the string '-c'.

지금은 FLAGS가 글로벌하게 선언이 되어있어서 argv가 전달된다는게 큰의미는 없다.
(전달안되도 어차피 글로벌이라 그냥 사용가능.tf.app.run()은 그냥 one line fast argument parser로 생각하면 될 듯 하다.)
"""

def main(argv=None):  # pylint: disable=unused-argument
    print(argv)
    """
    설명3:
    1) gfile은 약간 폴더,파일쪽 다루는 패키지인가보다, 만약 checkpoint_dir에서 지정 된 폴더가 없으면 그걸 만들어서
    2) os.path.join은 단순히 경로 만들어주는 함수다, 저절로 / 를 추가해주는게 편리한 점
    3) assert FLAGS.model로 명시한 모델이 있으면 통과 없으면 뒤에 있는 오류메시지 발생
    4) 해당 파이썬 파일을(인풋1) 인풋2로 복사한다.
    5) 해당 모델을 import한다, importlib.import_module()함수는 코드 과정 중에 패키지를 import 할 때 쓰는 듯 하다.
    6) data = get_data_provider(FLAGS.dataset, training=True) 에서 training에 따라 trainset 혹은 testset을 불러온다.
    7) 위에서 정의한 train함수를 실행 - train을 진행하기 전에 data.py파일을 살펴보자
    """
    if not gfile.Exists(FLAGS.checkpoint_dir):
        # gfile.DeleteRecursively(FLAGS.checkpoint_dir)
        gfile.MakeDirs(FLAGS.checkpoint_dir)
        model_file = os.path.join('models', FLAGS.model + '.py')
        assert gfile.Exists(model_file), 'no model file named: ' + model_file
        gfile.Copy(model_file, FLAGS.checkpoint_dir + '/save_model.py')
        gfile.Copy('./main.py',FLAGS.checkpoint_dir+'/save_main.py')
        gfile.Copy('./evaluate.py',FLAGS.checkpoint_dir+'/save_evaluate.py')
        gfile.Copy('./nnUtils.py',FLAGS.checkpoint_dir+'/save_nnUtils.py')
        gfile.Copy('./data.py',FLAGS.checkpoint_dir+'/save_data.py')
    m = importlib.import_module('models.' + FLAGS.model)
    data = get_data_provider(FLAGS.dataset, training=True)

    train(m.model, data,
          batch_size=FLAGS.batch_size,
          checkpoint_dir=FLAGS.checkpoint_dir,
          log_dir=FLAGS.log_dir,
          num_epochs=FLAGS.num_epochs)

"""

설명1:여기서부터 설명한다,일단 이 코드를 실행하면 위에서부터 라인바이라인 실행이 된다.
다만 함수는 라인바이라인 실행 되는 것이 아니라 함수가 있다는 것을 알려줄 수 있는 '선언'만 하고 넘어가게 된다.
그러면 마지막에 실행되는 것이 아래의 조건문이다.
모든 파일은 실행 될 때 __name__이 부여되는 듯 하다. __main__이라는건 이 파일이! 모듈 같이 간접적으로 실행되었다는게 아니라
직접적으로 실행되었다는 뜻이다, 즉 아래의 코드는 이 파일이 직접 실행 된건지, 아니면 다른 코드에 종속적으로 실행된건지를 판단한다.
-여기서 종속적으로 실행되면 __name__=? 인지는 아직 모른다.
이걸 판단해서 직접적으로 실행된거면 tf.app.run()을 실행한다, 아니면 아무것도 실행안된다, 즉 아무것도 안된다.
이는 이 파일이 부정할 수 없는 메인파일이라는 것을 말해준다.

여기서 처음에 들었던 의문-tf.app.run()은 argv를 받아서 main()이라는 이름의 함수(default)에 전달해서 실행시킨다고 하는데
왜 바로 아래 판단문에 그 함수의 내용을 쓰지 않는걸까? 굳이 한번의 매개를 거치는 이유가 무엇인가?
->답변: 우리는 FLAGS를 이용해서 커맨드창에서 파라미터설정을 입력 받는다, 그리고 그걸 wraping해서 함수에 쓰는거다.
tf.app.run()은 그걸 wraping해주는 좋은 함수이고, 그걸 main()함수에 전달하도록 되어있다. 즉 tf.app.run()함수 자체가 main()함수를
필요로 하는 것이다.
#Runs the program with an optional 'main' function and 'argv' list.
->그렇다면 질문은 tf.app.run()가 왜 이렇게 쓰이냐!로 되는데 그건 나중에 생각해보자.
추측: 만약 이 함수를 안쓰면 우리는 string인 cmd 명령어를 쪼개서, 그 중에 어떤 부분이 파일에 정의되어있는지 체크해서
그걸 argv로 만들어줘야한다. 
  nmt_parser = argparse.ArgumentParser()
  add_arguments(nmt_parser)
  FLAGS, unparsed = nmt_parser.parse_known_args()
  tf.app.run(main=main, argv=[sys.argv[0]] + unparsed)        #https://stackoverflow.com/questions/33703624/how-does-tf-app-run-work
보통 위처럼 세줄정도는 나올텐데, 그것보다는 한줄로 써서 위의 기능들을 취하는 것 아닐까?
그리고 main()으로 구분하면 가독성도 좋아지고, __main__이 아니어도 어떻게 접근이 가능할수도 있고..여러가지 가능성이 나오게되는 장점도 있다

"""
if __name__ == '__main__':
    # print(callable(main))
    # print(locals())
    tf.app.run()