from Myoption import *
import tensorflow as tf
import numpy as np
import customfunc
from openpyxl import load_workbook
FLAGS = tf.app.flags.FLAGS
@tf.RegisterGradient("fluc_grad")
def fluc_grad(op,grad):
    shape=op.inputs[1]._shape_as_list()
    return grad,tf.zeros(shape=shape)

def scalebybits(bits):
    return 2.0 ** (bits - 1)

def shiftbyval(x):
    return 2.0 ** tf.round(tf.log(x+1e-12) / tf.log(2.0))

def quantize(x, target_level=2 ** 8):
    assert np.floor(target_level) == target_level, "Target_level should be a integer"
    bits = np.ceil(np.log(target_level) / np.log(2.0))
    SCALE = scalebybits(bits)

    # shift=shiftbyval(tf.abs(x))  #We don't shift x in this function, this function is only for quantizatizing any input.
    if bits > 15:
        print("The target_level is to high, We don't quantize this.")
        # 15비트 넘어가면 양자화 안하겠다는건데, 나름 괜찮은 선택 같다, 그 위는 양자화해도 크게 득도 없고 아래에서 결판나니까
        y = x
    elif bits == 1:  # BNN
        print("The target_level=1 and we use BNN")
        y = tf.sign(x)
    elif target_level % 2 == 1:
        print('target_level is a odd number')
        y = tf.round(x * tf.floor(target_level / 2)) / SCALE
        # 20180410
        # 레벨 수는 맞추되, 그 레벨의 값이 2의 승수로 만든다.
        # 다만 이 부분에서 확인이 안된건 weight 값에 따라 학습결과가 달라지나 안달라지나를 확인해야한다.
        # -0.5, 0, 0.5와 -1, 0, 1의 학습결과가 달라지나 안달라지나?
        # BatchNorm쓰면 별로 안달리질거고, WAGE에서는 scale layer를 쓰는데 그걸 여기 구현하지는 않았다.
    else:
        print('target_level is a even number')
        SCALE = SCALE * 2
        y = (tf.sign(x)*(2*tf.ceil(tf.abs(x)*(target_level/2))-1)+
             tf.to_float(tf.equal(x,0))*tf.sign(tf.random_uniform(x.get_shape(),-1.,1.)))/SCALE
    return tf.stop_gradient(y - x) + x

def clip(x, target_level=2 ** 8):
    # 이게 왜 필요하냐?
    # 레벨 수가 몇이든 간에 결국 해당하는 bits의 2의 승수배로 weight 값을 주게된다.
    # 그니까 그 최댓값을 넘어가는 값은 다 최댓값으로 맵핑시킬 필요가 있다.
    if target_level==None:
        MAX=1.
        MIN=-1.

    else:
        assert np.floor(target_level) == target_level, "Target_level should be a integer"
        bits = np.ceil(np.log(target_level) / np.log(2.0))
        SCALE = scalebybits(bits)
        if bits > 15 or bits == 1:
            delta = 0.
        elif target_level % 2 == 1:
            limit = SCALE - np.floor(target_level/2)
            delta = 1. / SCALE
        else:
            limit = SCALE*2 - target_level + 1
            delta = 1. / (SCALE*2)
        MAX = +1 - limit * delta
        MIN = -1 + limit * delta
    y = tf.clip_by_value(x, MIN, MAX, name='saturate')
    return tf.stop_gradient(y - x) + x

def quantize_G(x, target_level=2 ** 8):
    bitsG = np.ceil(np.log(target_level) / np.log(2.0))
    SCALE = scalebybits(bitsG)
    with tf.name_scope('Quantize_G'):
        if bitsG > 15:
            return x
        else:
            if x.name.lower().find('batchnorm') > -1:
                return x  # batch norm parameters, not quantize now
            xmax = tf.reduce_max(tf.abs(x))
            tf.add_to_collection('testt',xmax)
            y = x / shiftbyval(xmax)
            norm = quantize(LR * y, 4095)
            norm_sign = tf.sign(norm)
            norm_abs = tf.abs(norm)
            norm_int = tf.floor(norm_abs)
            norm_float = norm_abs - norm_int
            rand_float = tf.random_uniform(x.get_shape(), 0, 1)
            norm = norm_sign * (norm_int + 0.5 * (tf.sign(norm_float - rand_float) + 1))
            return norm / SCALE

def quantize_W(x, target_level=2 ** 8):
    bitsW = np.ceil(np.log(target_level) / np.log(2.0))
    with tf.name_scope('Quantize_W'):
        if bitsW > 15:
            return x
        else:
            xmax = tf.to_float(tf.reduce_max(tf.abs(x)))
            y = x / shiftbyval(xmax)
            #y = clip(quantize(y, target_level), target_level)
            y = quantize(clip(y, None), target_level)
            # 20180410: clipr과 quantize의 적용 순서에 대한 명확한 이해가 없다.
            # we scale W in QW rather than QA for simplicity
            return x + tf.stop_gradient(y - x)  # skip derivation of Quantize and Clip
def quantize_old(x,target_level=FLAGS.Wq_target_level):
    assert (target_level>1) and (int(target_level)==target_level), \
        'target_level is not valid, please input the integer bigger than 1 '
    """
    Clip and binarize tensor using the straight through estimator (STE) for the gradient.
    """
    g = tf.get_default_graph()
    if target_level==2:
        with tf.name_scope("Binarized") as name:
            with g.gradient_override_map({"Sign": "Identity"}):
                x=tf.clip_by_value(x,-1,1)
                results=tf.sign(x)
                if FLAGS.bit_deterministic:
                    results=results+tf.cast(x>(tf.reduce_max(x)/2),dtype=tf.float32)*0.4
    else:
        with g.gradient_override_map({"Round":"Identity"}):
            x = tf.tanh(x)
            #####################################################
            #1:
            x = x / tf.reduce_max(tf.abs(x)) * 0.5 + 0.5
            #2:
            # x = x / tf.reduce_mean(tf.abs(x))
            # x = tf.clip_by_value(x,-1,1) * 0.5 + 0.5
            #####################################################
            x=tf.round(x*(target_level-1))/(target_level-1)
        results=2*x-1
    return results

def memory_mode(x,choice=0):
    #기본정보, 신경 쓸 필요 없음
    tf.add_to_collection("Original_Weight", x)  # stage1
    tf.add_to_collection("Drift_value", tf.convert_to_tensor([]))
    tf.add_to_collection("Drift_step", tf.convert_to_tensor([]))
    filter_shape = x.get_shape().as_list()
    target_level = FLAGS.W_target_level
    bits = np.ceil(np.log(target_level) / np.log(2.0))
    variation=FLAGS.Variation

    if target_level%2==1:
        level_index = tf.to_int32(tf.round(x*scalebybits(bits)+tf.ceil(target_level/2)))-1
    else:
        level_index = tf.to_int32(tf.ceil((x*scalebybits(bits)*2+target_level)/2))-1

    #정보 from the excel file
    wb = load_workbook(filename='level_info.xlsx')
    sheetname=str(target_level)+'Levels'
    if str(wb.sheetnames).find(sheetname)>-1:
        ws = wb[sheetname]
        r = np.array([cell.value for cell in tuple(ws.rows)[3 * choice]])
        r = np.array([simple_r for simple_r in range(1,target_level+1)]).astype('float32') \
            if r[0]==None else r.astype('float32')
        r_var = np.array([cell.value for cell in tuple(ws.rows)[3 * choice + 1]]).astype('float32') \
            if variation else np.zeros([target_level]).astype('float32')
        r_ref = [cell.value for cell in tuple(ws.rows)[3 * choice + 2]][1:] if variation else (r[1:]+r[0:-1])/2
        # ref에 variation 안넣었다.
    else:
        r=np.array([r for r in range(1,target_level+1)]).astype('float32')
        r_var=np.zeros([target_level]).astype('float32')
        r_ref=(r[1:]+r[0:-1])/2
    # 준비
    tr = tf.gather(r, level_index)
    tr_var = tf.gather(r_var, level_index)
    tf.add_to_collection('Binarized_Weight', tr)  # stage2/target_resistance

    if FLAGS.Variation==True:
        pre_tr = \
            tf.Variable(initial_value=tf.ones(shape=filter_shape)*r[0],name='pre_target_resistance',trainable=False)
        pre_tr_place = tf.placeholder(dtype=tf.float32, shape=filter_shape)
        pre_tr_update_op = pre_tr.assign(pre_tr_place)
        pre_wr = \
            tf.Variable(initial_value=tf.ones(shape=filter_shape)*r[0], name='pre_written_resistance', trainable=False)
        pre_wr_place = tf.placeholder(dtype=tf.float32, shape=filter_shape)
        pre_wr_update_op = pre_wr.assign(pre_wr_place)
        tf.add_to_collection('pre_Wbin',pre_tr_place)
        tf.add_to_collection('pre_Wbin_update_op', pre_tr_update_op)
        tf.add_to_collection('pre_Wfluc', pre_wr_place)
        tf.add_to_collection('pre_Wfluc_update_op', pre_wr_update_op)
        # 왜  바로 pre_r_place로 비교하지않는가,매번 feed 해주는게  그렇게  어려운 일인가?
        # write phase
        fluc_value=tf.reshape(tf.distributions.Normal(loc=0.,scale=tr_var).sample(1),shape=filter_shape)
    # if Memristor resistance를 read해서 write여부를 결정할거면-오류수정
    # pre_wr_index=customfunc.make_index(pre_wr,r_ref)
    # tr_index = customfunc.make_index(tr, r_ref)
    # keep_bool=tf.to_float(tf.equal(pre_wr_index,tr_index))
    # if pre-iteration의 target resistance를 기준으로 하여 이번 cycle의 target resistance가 변했는지 여부로 결정할거면,오류무시
        keep_bool=tf.to_float(tf.equal(pre_tr,tr))
        update_bool=1-keep_bool
        wr=keep_bool*pre_wr+update_bool*(tr+fluc_value)
    else:
        wr=tr
    tf.add_to_collection('Fluctuated_Weight', wr) # stage3
    # read phase
    rr=wr
    # rr=tf.to_float(customfunc.make_index(wr,r_ref))
    # if target_level%2==1:
    #     rr = (rr-tf.ceil(target_level/2))/scalebybits(bits)
    # else:
    #     rr = (2*rr-1-target_level)/scalebybits(bits+1)
    # tf.add_to_collection('Read_Weight', rr) # stage4
    return rr

def fluctuate(x,scale=1,target_level=FLAGS.Wq_target_level,Drift=False):
    filter_shape = x.get_shape().as_list()
    pre_Wbin = tf.Variable(initial_value=tf.zeros(shape=filter_shape),name='pre_Wbin',trainable=False)
    pre_Wbin_val_place=tf.placeholder(dtype=tf.float32,shape=filter_shape)
    pre_Wbin_update_op=pre_Wbin.assign(pre_Wbin_val_place)
    pre_Wfluc = tf.Variable(initial_value=tf.zeros(shape=filter_shape),name='pre_Wfluc',trainable=False)
    pre_Wfluc_val_place = tf.placeholder(dtype=tf.float32, shape=filter_shape)
    pre_Wfluc_update_op = pre_Wfluc.assign(pre_Wfluc_val_place)
    tf.add_to_collection('pre_Wbin',pre_Wbin_val_place)
    tf.add_to_collection('pre_Wbin_update_op', pre_Wbin_update_op)
    tf.add_to_collection('pre_Wfluc', pre_Wfluc_val_place)
    tf.add_to_collection('pre_Wfluc_update_op', pre_Wfluc_update_op)

    if FLAGS.Variation==True:
        # Reset_Meanmean, Reset_Meanstd = 0., 0.1707
        # Reset_Stdmean, Reset_Stdstd = 0.0942, 0.01884
        # Set_Meanmean, Set_Meanstd = 0., 0.1538
        # Set_Stdmean, Set_Stdstd = 0.1311, 0.06894
        Reset_Meanmean, Reset_Meanstd = 0., 0.0000000001
        Reset_Stdmean, Reset_Stdstd = 0., 0.
        Set_Meanmean, Set_Meanstd = 0., 0.0000000001
        Set_Stdmean, Set_Stdstd = 0., 0.
        # 퍼센테이지 맵핑으로 간단하게 구했다,정밀하지는 않다.
    else:
        Reset_Meanmean, Reset_Meanstd = 0., 0.0000000001
        Reset_Stdmean, Reset_Stdstd = 0., 0.
        Set_Meanmean, Set_Meanstd = 0., 0.0000000001
        Set_Stdmean, Set_Stdstd = 0., 0.

    with tf.name_scope("Fluctuated") as name:
        # variation을 만들자
        fluc_list=[]
        # for i in range(2):
        #      fluc_list += [customfunc.get_distrib([0., 0., FLAGS.target_std, 0.04],filter_shape)]
        fluc_list += [customfunc.get_distrib([Set_Meanmean, Set_Meanstd, Set_Stdmean, Set_Stdstd], filter_shape)]
        fluc_list += [customfunc.get_distrib([Reset_Meanmean, Reset_Meanstd, Reset_Stdmean, Reset_Stdstd], filter_shape)]
        # 여기서 weight value에 따라 선택적으로 fluc값을 generate하면 random값이 훨씬 줄어들겠다.
        # assign 1 to elements which have same state with pre-state
        keep_element = tf.cast(tf.equal(x,pre_Wbin), tf.float32)
        # assign 1 to elements which have different state with pre-state
        update_element = tf.cast(tf.not_equal(x,pre_Wbin), tf.float32)
        # 이 부분에서 쓰지도않는 랜덤 값이 많이 발생하는데 일단 두고 나중에 고치든가 하자
        Wfluc_Reset = update_element * fluc_list[-1] * tf.cast(x > 0, tf.float32)
        Wfluc_Set = update_element * fluc_list[0] * tf.cast(x <= 0, tf.float32)
        # Variation으로 인한 delta

        step_col = tf.get_collection("Step")
        if Drift and step_col!=[]:   #drift의 구성요소, d value and step_num을 다뤄주면 된다.
            # 1:d value를 만들자.
            drift_value = np.array(np.ones(shape=filter_shape) * 0.09)  #drift_value가 dvalue이다,즉 power law에서 지수부분
            drift_Meanmean, drift_Meanstd = 0.09, 0.001
            drift_Stdmean, drift_Stdstd = 0, 0.0003
            new_value=customfunc.get_distrib([drift_Meanmean, drift_Meanstd, drift_Stdmean,drift_Stdstd], filter_shape)
            new_drift = tf.cast((pre_Wbin <= 0), dtype=tf.float32)* tf.cast((x > 0), dtype=tf.float32)
            # 전사이클에서는 -1이었고 이번 사이클에 1로 업데이트 된 element 부분을 1로, 즉 한번 set->reset이 될 때마다 d value가 바뀐다.
            # note1:사실 매 iteration drift 할 때마다 d value조차도 약간 바뀌지만 그건 일단 고려하지 않았다.
            # note2:여기서의 drift는 순전히 2levels일 때만 고려,그 이상의 MLC에서는 drift를 사용하지않기를 추천(unrealistic).
            drift_value = new_value * new_drift + drift_value * (1 - new_drift)
            tf.add_to_collection("Drift_value",drift_value)

            step = tf.Variable(tf.zeros(shape=filter_shape, dtype=tf.float32), trainable=False)
            step = tf.assign(step, keep_element*(step+1),name="Drift_step")
            tf.add_to_collection("Drift_step",step)
            # step은 각 웨이트가 드리프트를 지금 몇스텝째 하고있는지를 나타낸다.

            drift_factor = (step+1.)/(tf.cast(tf.equal(step,0.),dtype=tf.float32)+step)
            # step=0인 부분은 결국 분모=1+0=1, step!=0인 부분은 분모=0+step=step
            drift_scale =  (tf.log(drift_factor) / tf.log(10.))*drift_value
            # log가 들어가는 이유는 이 코드 작성 당시 R->W의 맵핑을 log함수라고 생각했었기 때문이다. 지금은 로그맵핑을 쓰지는않을 것 같아서
            # MAC에서 맵핑을 어떻게 하느냐에 따라 수정이 필요할거 같다.
        else:
            tf.add_to_collection("Drift_value", tf.convert_to_tensor([]))
            tf.add_to_collection("Drift_step", tf.convert_to_tensor([])) #이 두줄은 summary부분 쉽게 만드려고 했던거다.
            drift_scale = tf.constant(0.)
        g = tf.get_default_graph()
        with g.gradient_override_map({"Mul": "fluc_grad", "Cast": "Identity",
                                      "Equal": "fluc_grad", "Greater": "fluc_grad",
                                      "LessEqual": "fluc_grad", "NotEqual": "fluc_grad",
                                      "Add": "fluc_grad"}):
            with tf.control_dependencies([drift_scale]):
                Wfluc = tf.multiply(x, update_element) +tf.cast(tf.greater(pre_Wbin,0), tf.float32) * keep_element * (pre_Wfluc + drift_scale)+ \
                        tf.cast(tf.less_equal(pre_Wbin,0), tf.float32) * keep_element * pre_Wfluc * 1. \
                        + Wfluc_Reset + Wfluc_Set
                #variation drift 각각  출력을  만드는게 더이쁘겠다.  그리고  둘  다 안하면  이  파트어떻게되는거야
        return Wfluc

def BinarizedWeightOnlyAffine(nOutputPlane, bias=True, name=None, reuse=None, Drift=False):
    def bwo_affineLayer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=name, default_name='Affine', reuse=reuse):

            '''
            Note that we use binarized version of the input (bin_x) and the weights (bin_w). Since the binarized function uses STE
            we calculate the gradients using bin_x and bin_w but we update w (the full precition version).
            '''
            reshaped = tf.reshape(x, [x.get_shape().as_list()[0], -1])
            nInputPlane = reshaped.get_shape().as_list()[1]
            w = tf.get_variable('weight', [nInputPlane, nOutputPlane], initializer=tf.contrib.layers.xavier_initializer())
            # read_w=tf.stop_gradient(memory_mode(w, choice=0) - w) + w
            quantized_w = quantize_W(w,FLAGS.Wq_target_level)
            bin_w = quantize_old(w)
            fluc_w = fluctuate(bin_w,Drift=Drift)
            tf.add_to_collection('Original_Weight', w)
            tf.add_to_collection('Binarized_Weight', bin_w)
            tf.add_to_collection('Fluctuated_Weight', fluc_w)
            # tf.add_to_collection('Original_Weight',w)
            # tf.add_to_collection("Drift_value", tf.convert_to_tensor([]))
            # tf.add_to_collection("Drift_step", tf.convert_to_tensor([]))
            output = tf.matmul(reshaped, fluc_w)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                output = tf.nn.bias_add(output, b)
        return output
    return bwo_affineLayer
"""
주의:binarize(x)가 activation까지 바이너리화 하는 중이다.
아래의 세개는 각각 Binary Conv layer,Binary Conv layer for weight, Vanilla Conv layer
"""
def BinarizedSpatialConvolution(nOutputPlane, kW, kH, dW=1, dH=1,
        padding='VALID', bias=True, reuse=None, name='BinarizedSpatialConvolution',bin=True,fluc=True,Drift=False):
    def b_conv2d(x, is_training=True):
        nInputPlane = x.get_shape().as_list()[3]
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name,reuse=reuse):
            w = tf.get_variable('weight', [kH, kW, nInputPlane, nOutputPlane],
                            initializer=tf.contrib.layers.xavier_initializer_conv2d())
            bin_x = quantize(x) if bin else x
            tf.add_to_collection(tf.GraphKeys.ACTIVATIONS, bin_x)
            bin_w = quantize(w)
            fluc_w = fluctuate(bin_w,Drift=Drift) if fluc else bin_w
            tf.add_to_collection('Original_Weight',w)
            tf.add_to_collection('Binarized_Weight', bin_w)
            tf.add_to_collection('Fluctuated_Weight', fluc_w)
            '''
            Note that we use binarized version of the input and the weights. Since the binarized function uses STE
            we calculate the gradients using bin_x and bin_w but we update w (the full precition version).
            '''
            out = tf.nn.conv2d(bin_x, fluc_w, strides=[1, dH, dW, 1], padding=padding)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                out = tf.nn.bias_add(out, b)
            return out
    return b_conv2d

def BinarizedWeightOnlySpatialConvolution(nOutputPlane, kW, kH, dW=1, dH=1,
        padding='VALID', bias=True, reuse=None, name='BinarizedWeightOnlySpatialConvolution',fluc=True,Drift=False):
    '''
    This function is used only at the first layer of the model as we dont want to binarized the RGB images
    '''
    def bc_conv2d(x, is_training=True):
        nInputPlane = x.get_shape().as_list()[3]
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name, reuse=reuse):
            w = tf.get_variable('weight', [kH, kW, nInputPlane, nOutputPlane],
                            initializer=tf.contrib.layers.xavier_initializer_conv2d())
            bin_w = quantize(w)
            fluc_w = fluctuate(bin_w,Drift=Drift) if fluc else bin_w
            tf.add_to_collection('Original_Weight', w)
            tf.add_to_collection('Binarized_Weight', bin_w)
            tf.add_to_collection('Fluctuated_Weight', fluc_w)

            out = tf.nn.conv2d(x, fluc_w, strides=[1, dH, dW, 1], padding=padding)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                out = tf.nn.bias_add(out, b)
            return out
    return bc_conv2d

def SpatialConvolution(nOutputPlane, kW, kH, dW=1, dH=1,
        padding='VALID', bias=True, reuse=None, name='SpatialConvolution'):
    def conv2d(x, is_training=True):
        nInputPlane = x.get_shape().as_list()[3]
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name, reuse=reuse):
            w = tf.get_variable('weight', [kH, kW, nInputPlane, nOutputPlane],
                            initializer=tf.contrib.layers.xavier_initializer_conv2d())
            tf.add_to_collection('Original_Weight', w)
            out = tf.nn.conv2d(x, w, strides=[1, dH, dW, 1], padding=padding)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                out = tf.nn.bias_add(out, b)
            return out
    return conv2d
"""
아래의 세개는 각각 Vanilla Affine layer,Binary Affine layer,Binary Affine layer for weight, 
"""
#Fully connected layer
def Affine(nOutputPlane, bias=True, name=None, reuse=None):
    def affineLayer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=name, default_name='Affine', reuse=reuse):
            temp=x.get_shape().as_list()
            reshaped = tf.reshape(x, [-1,np.array(temp[1:]).prod()])
            nInputPlane = reshaped.get_shape().as_list()[1]
            w = tf.get_variable('weight', [nInputPlane, nOutputPlane], initializer=tf.contrib.layers.xavier_initializer())
            tf.add_to_collection('Original_Weight', w)
            output = tf.matmul(reshaped, w)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                output = tf.nn.bias_add(output, b)
        return output
    return affineLayer

def BinarizedAffine(nOutputPlane, bias=True, name=None, reuse=None,bin=True,fluc=True,Drift=False):
    def b_affineLayer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=name, default_name='Affine', reuse=reuse):
            '''
            Note that we use binarized version of the input (bin_x) and the weights (bin_w). Since the binarized function uses STE
            we calculate the gradients using bin_x and bin_w but we update w (the full precition version).
            '''
            bin_x = quantize(x) if bin else x
            tf.add_to_collection(tf.GraphKeys.ACTIVATIONS, bin_x)
            reshaped = tf.reshape(bin_x, [x.get_shape().as_list()[0], -1])
            nInputPlane = reshaped.get_shape().as_list()[1]
            w = tf.get_variable('weight', [nInputPlane, nOutputPlane], initializer=tf.contrib.layers.xavier_initializer())
            bin_w = quantize(w)
            fluc_w = fluctuate(bin_w,Drift=Drift) if fluc else bin_w
            tf.add_to_collection('Original_Weight', w)
            tf.add_to_collection('Binarized_Weight', bin_w)
            tf.add_to_collection('Fluctuated_Weight', fluc_w)

            output = tf.matmul(reshaped, fluc_w)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                output = tf.nn.bias_add(output, b)
        return output
    return b_affineLayer



#bias 더 해주는 레이어
def Linear(nInputPlane, nOutputPlane):
    return Affine(nInputPlane, nOutputPlane, add_bias=False)

#여기서는 Sequential을 구성하기 위해서 일반레이어함수를 Sequential용으로 바꿔준다.
#*args,**kwargs 이 둘은 지금은 확정되지않았지만, 추후에 들어올 수 있는 인풋을 담당한다.
def wrapNN(f,*args,**kwargs):
    def layer(x, is_training=True):
        return f(x,*args,**kwargs)
    return layer

#이것도 사실 위의 함수와 기능은 똑같으나, dropout이기 때문에 training일 때와 test일 때를 다르게 해줘야해서, 따로 함수를 정의
def Dropout(p, name='Dropout'):
    def dropout_layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            # def drop(): return tf.nn.dropout(x,p)
            # def no_drop(): return x
            # return tf.cond(is_training, drop, no_drop)
            if is_training:
                return tf.nn.dropout(x,p)
            else:
                return x
    return dropout_layer

"""
아래의 두 함수는 activation이다, activation layer를 만드는 셈
다시 한번 말하지만 아래처럼 함수로 반환하는 이유는 인풋의 자유를 남겨놓기 위해서다.
HardTanh은 Tanh랑 다르다.
"""
def ReLU(name='ReLU'):
    def layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            return tf.nn.relu(x)
    return layer

def HardTanh(name='HardTanh'):
    def layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            return tf.clip_by_value(x,-1,1)
    return layer


def View(shape, name='View'):
    return wrapNN(tf.reshape,shape=shape)
"""
    kH = kH or kW    의 뜻은 전자가 없으면 후자로 쓴다는 뜻
->최소 kW는 인풋으로 넣어줘야 멀쩡한 함수 실행이 가능하다,
kW는 window size, kH없으면 윈도우는 kH=kW인 정사각형이 된다 
만약 strides도 안넣어준다면 stride는 윈도우의 가로세로와 같다.
Tip:파이썬에서는 and와 or이 본래 논리연산자의 기능을 하면서 인풋으로 들어간 값들을 이용 할 수 있도록하여
효율을 극대화 하였다.
"""
def SpatialMaxPooling(kW, kH=None, dW=None, dH=None, padding='VALID',
            name='SpatialMaxPooling'):
    kH = kH or kW
    dW = dW or kW
    dH = dH or kH
    def max_pool(x,is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
              return tf.nn.max_pool(x, ksize=[1, kW, kH, 1], strides=[1, dW, dH, 1], padding=padding)
    return max_pool

def SpatialAveragePooling(kW, kH=None, dW=None, dH=None, padding='VALID',
        name='SpatialAveragePooling'):
    kH = kH or kW
    dW = dW or kW
    dH = dH or kH
    def avg_pool(x,is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
              return tf.nn.avg_pool(x, ksize=[1, kW, kH, 1], strides=[1, dW, dH, 1], padding=padding)
    return avg_pool

def BatchNormalization(name='BatchNormalization',*kargs, **kwargs):
    with tf.variable_scope(name_or_scope=name):
        output=wrapNN(tf.contrib.layers.batch_norm, *kargs, **kwargs)
    return output

def Sequential(moduleList):
    def model(x, is_training=True):
    # Create model
        output = x
        #with tf.variable_op_scope([x], None, name):
        for i,m in enumerate(moduleList):
            output = m(output, is_training=is_training)
            tf.add_to_collection(tf.GraphKeys.ACTIVATIONS, output)
        return output
    return model
"""
아래 두 함수는 현재 사용하지 않음, 다만 나중에 네트워크 바꿀 때 사용할 수도 있어서 냅두었음
"""
#병렬연결모드
def Concat(moduleList, dim=3):
    def model(x, is_training=True):
    # Create model
        outputs = []
        for i,m in enumerate(moduleList):
            name = 'layer_'+str(i)
            with tf.variable_scope(values=[x], name_or_scope=name, default_name='Layer'):
                outputs[i] = m(x, is_training=is_training)
            output = tf.concat(dim, outputs)
        return output
    return model

#말그대로 residual
def Residual(moduleList, name='Residual'):
    m = Sequential(moduleList)
    def model(x, is_training=True):
    # Create model
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            output = tf.add(m(x, is_training=is_training), x)
            return output
    return model
