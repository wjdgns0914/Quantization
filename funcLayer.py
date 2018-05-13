from params import *
import tensorflow as tf
import numpy as np
import funcQuantize
from openpyxl import load_workbook
FLAGS = tf.app.flags.FLAGS
W_level=FLAGS.W_target_level
Wq_level=FLAGS.Wq_target_level
choice=FLAGS.choice
Inter=eval('['+FLAGS.Inter_variation_options+']')
Intra=eval('['+FLAGS.Intra_variation_options+']')
mean_a = 3.156e+06 * 0.9996
mean_b = 4.463 * 1.0647
mean_c = 2.713 * 1.0004
mean_d = -2.509e+04 * 1.0368
# 정보 from the excel file
wb = load_workbook(filename='level_info.xlsx')
sheetname = str(W_level) + 'Levels'
# 오늘: 이 부분 훨씬 짧게 만들 수 있다, none인거 검사하고 시작하면 됨
# 이 부분은 비록 보기에는 지저분해도 속도에는 별영향을 주지 않는다.
if (Intra[0] or Inter[0]) and (str(wb.sheetnames).find(sheetname) > -1):
    ws = wb[sheetname]
    print("We will load writing information from %s"%sheetname)
    assert ws['A' + str(4 * choice + 1)] != None,"This sheet doesn't have voltage information"
    volt = np.array([cell.value for cell in tuple(ws.rows)[3 * choice]]).astype('float32')
    if ws['A' + str(4 * choice + 2)] != None:
        r = np.array([cell.value for cell in tuple(ws.rows)[3 * choice+1]]).astype('float32')
    if ws['A' + str(4 * choice + 3)] != None:
        r_std = np.array([cell.value for cell in tuple(ws.rows)[3 * choice + 2]]).astype('float32')
    if ws['B' + str(4 * choice + 4)] != None:
        r_ref = np.array([cell.value for cell in tuple(ws.rows)[3 * choice + 3][1:]]).astype('float32')
        r_ref = (r[1:] + r[0:-1]) / 2
    # elif FLAGS.Inter_variation_options_target:
    #     r_std = np.ones([W_level], dtype='float32') * FLAGS.target_std
    # else:
    #     r_std = np.zeros([W_level], dtype='float32')
    # if FLAGS.Load_ref and ws['A' + str(3 * choice + 3)] != None:  # ref에 variation 안넣었다.
    #     r_ref = np.array([cell.value for cell in tuple(ws.rows)[3 * choice + 3][1:]]).astype('float32')
    #     ########################################################################################################################################################
    # else:
    #     r_ref = (r[1:] + r[0:-1]) / 2
    r = mean_a / (1 + np.exp(-mean_b * (volt - mean_c))) - mean_d
    r_ref = (r[1:] + r[0:-1]) / 2
    # target resistance
    if Intra[0]:
        meanofstd = 0.6 * r_std
        stdofstd = 0.5 * meanofstd
        if Intra[1]:
            print("To be continued..Sorry!")
            exit(0)
        if Intra[3]:
            meanofstd = Intra[4] * meanofstd
            stdofstd = Intra[4] * stdofstd
    else:
        meanofstd = 0. * r_std
        stdofstd = 0. * meanofstd

    if Inter[0]:
        std_a = 3.156e+06 * 0.0698
        std_b = 4.463 * 0.1824
        std_c = 2.713 * 0.0354
        std_d = 2.509e+04 * 0.7708
        # stdofstd is free, 0일 수도 있고 아닐 수도 있다.
        if Inter[1]:
            print("To be continued..Sorry!")
            exit(0)
        if Inter[3]:
            std_a = Inter[4] * std_a
            std_b = Inter[4] * std_b
            std_c = Inter[4] * std_c
            std_d = Inter[4] * std_d
    else:
        std_a = 0.
        std_b = 0.
        std_c = 0.
        std_d = 0.
        stdofstd = 0. * meanofstd  # 무조건 0이어야한다.
else:
    print("There is no sheet named as %s" % sheetname)
    volt = np.linspace(1.5,4,W_level).astype('float32')
    r = mean_a / (1 + np.exp(-mean_b * (volt - mean_c))) - mean_d
    r_std = np.zeros([W_level], dtype='float32')
    r_ref = (r[1:] + r[0:-1]) / 2
    std_a = 0.
    std_b = 0.
    std_c = 0.
    std_d = 0.
    meanofstd = 0. * r_std
    stdofstd = 0. * meanofstd

print("Set")
#Fully connected layer
def BinarizedAffine(nOutputPlane, bias=True, name=None, reuse=None,bin=False,fluc=True,Drift=False):
    def b_affineLayer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=name, default_name='Affine', reuse=reuse):
            '''
            Note that we use binarized version of the input (bin_x) and the weights (bin_w). Since the binarized function uses STE
            we calculate the gradients using bin_x and bin_w but we update w (the full precition version).
            '''
            bin_x = funcQuantize.quantize_W(x, Wq_level) if bin else x  #원래 funcQ.quantize_A라는 함수를 써야함, 근데 activation일단 안건드리니까 이렇게 둠
            reshaped = tf.reshape(bin_x, [x.get_shape().as_list()[0], -1])
            nInputPlane = reshaped.get_shape().as_list()[1]
            w = tf.get_variable('0/Original_weight', [nInputPlane, nOutputPlane], initializer=tf.contrib.layers.xavier_initializer())
            filter_shape=w.get_shape().as_list()
            a = tf.Variable(np.random.normal(loc=mean_a, scale=std_a,size=filter_shape),trainable=False,dtype=tf.float32)
            b = tf.Variable(np.random.normal(loc=mean_b, scale=std_b,size=filter_shape),trainable=False,dtype=tf.float32)
            c = tf.Variable(np.random.normal(loc=mean_c, scale=std_c,size=filter_shape),trainable=False,dtype=tf.float32)
            d = tf.Variable(np.random.normal(loc=mean_d, scale=std_d,size=filter_shape),trainable=False,dtype=tf.float32)
            std_val=tf.Variable(np.random.normal(loc=meanofstd,scale=stdofstd,size=filter_shape+list(meanofstd.shape)),trainable=False,dtype=tf.float32)
            w_written = tf.assign(w, funcQuantize.write_memory(w,cell_info=[a,b,c,d,std_val],write_info=[volt,r,r_ref],target_level=W_level))  #1)w는 W_level로 양자화되어있다고 가정한다. 2)write_memory에서의 target_level은 w를 양자화한 target_level과 같아야한다.
            with tf.control_dependencies([w_written]):
                w_propagated = tf.identity(funcQuantize.quantize_W(w, target_level=Wq_level), name='4/Quantized_weight')
            tf.add_to_collection('Propagated_Weight',w_propagated)
            tf.add_to_collection(tf.GraphKeys.ACTIVATIONS, bin_x)
            output = tf.matmul(reshaped, w_propagated)
            if bias:
                b = tf.get_variable('bias', [nOutputPlane],initializer=tf.zeros_initializer)
                output = tf.nn.bias_add(output, b)
        return output
    return b_affineLayer
#여기서는 Sequential을 구성하기 위해서 일반레이어함수를 Sequential용으로 바꿔준다.
#*args,**kwargs 이 둘은 지금은 확정되지않았지만, 추후에 들어올 수 있는 인풋을 담당한다.
def wrapNN(f,*args,**kwargs):
    def layer(x, is_training=True):
        return f(x,*args,**kwargs)
    return layer

#이것도 사실 위의 함수와 기능은 똑같으나, dropout이기 때문에 training일 때와 test일 때를 다르게 해줘야해서, 따로 함수를 정의
def Dropout(p, name='Dropout'):
    def dropout_layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            # def drop(): return tf.nn.dropout(x,p)
            # def no_drop(): return x
            # return tf.cond(is_training, drop, no_drop)
            if is_training:
                return tf.nn.dropout(x,p)
            else:
                return x
    return dropout_layer

"""
아래의 두 함수는 activation이다, activation layer를 만드는 셈
다시 한번 말하지만 아래처럼 함수로 반환하는 이유는 인풋의 자유를 남겨놓기 위해서다.
HardTanh은 Tanh랑 다르다.
"""
def ReLU(name='ReLU'):
    def layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            return tf.nn.relu(x)
    return layer
def Sigmoid(name='Sigmoid'):
    def layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            return tf.nn.sigmoid(x)
    return layer

def HardTanh(name='HardTanh'):
    def layer(x, is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            return tf.clip_by_value(x,-1,1)
    return layer


def View(shape, name='View'):
    return wrapNN(tf.reshape,shape=shape)
"""
    kH = kH or kW    의 뜻은 전자가 없으면 후자로 쓴다는 뜻
->최소 kW는 인풋으로 넣어줘야 멀쩡한 함수 실행이 가능하다,
kW는 window size, kH없으면 윈도우는 kH=kW인 정사각형이 된다 
만약 strides도 안넣어준다면 stride는 윈도우의 가로세로와 같다.
Tip:파이썬에서는 and와 or이 본래 논리연산자의 기능을 하면서 인풋으로 들어간 값들을 이용 할 수 있도록하여
효율을 극대화 하였다.
"""
def SpatialMaxPooling(kW, kH=None, dW=None, dH=None, padding='VALID',
            name='SpatialMaxPooling'):
    kH = kH or kW
    dW = dW or kW
    dH = dH or kH
    def max_pool(x,is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
              return tf.nn.max_pool(x, ksize=[1, kW, kH, 1], strides=[1, dW, dH, 1], padding=padding)
    return max_pool

def SpatialAveragePooling(kW, kH=None, dW=None, dH=None, padding='VALID',
        name='SpatialAveragePooling'):
    kH = kH or kW
    dW = dW or kW
    dH = dH or kH
    def avg_pool(x,is_training=True):
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
              return tf.nn.avg_pool(x, ksize=[1, kW, kH, 1], strides=[1, dW, dH, 1], padding=padding)
    return avg_pool

def BatchNormalization(name='BatchNormalization',*kargs, **kwargs):
    with tf.variable_scope(name_or_scope=name):
        output=wrapNN(tf.contrib.layers.batch_norm, *kargs, **kwargs)
    return output

def Sequential(moduleList):
    def model(x, is_training=True):
    # Create model
        output = x
        #with tf.variable_op_scope([x], None, name):
        for i,m in enumerate(moduleList):
            output = m(output, is_training=is_training)
            tf.add_to_collection(tf.GraphKeys.ACTIVATIONS, output)
        return output
    return model
"""
아래 두 함수는 현재 사용하지 않음, 다만 나중에 네트워크 바꿀 때 사용할 수도 있어서 냅두었음
"""
#병렬연결모드
def Concat(moduleList, dim=3):
    def model(x, is_training=True):
    # Create model
        outputs = []
        for i,m in enumerate(moduleList):
            name = 'layer_'+str(i)
            with tf.variable_scope(values=[x], name_or_scope=name, default_name='Layer'):
                outputs[i] = m(x, is_training=is_training)
            output = tf.concat(dim, outputs)
        return output
    return model

#말그대로 residual
def Residual(moduleList, name='Residual'):
    m = Sequential(moduleList)
    def model(x, is_training=True):
    # Create model
        with tf.variable_scope(values=[x], name_or_scope=None, default_name=name):
            output = tf.add(m(x, is_training=is_training), x)
            return output
    return model
